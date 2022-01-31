from openpyxl import Workbook, load_workbook, workbook
from openpyxl.worksheet import worksheet
from alum import Alum

def asking_info():
    component = input()
    
def main():

    alum_file = load_workbook(filename="Alumni Records.xlsx")
    # print(alum_file.sheetnames)
    #Gets the top sheet and it's title
    master = input("What file are you updating?: ")
    master_sheet = alum_file[master]

    update_num = int(input("How many profiles are you updating?"))

    for profile in update_num:
        alumni = Alum()


   

    alum_file.save(filename="Alumni Records.xlsx")
    print("done")
    #identifying headers
  
main()