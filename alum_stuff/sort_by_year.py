from openpyxl import Workbook, load_workbook, workbook
from openpyxl.worksheet import worksheet

def sort(file, master_sheet):
    row = 2
    for i in range(2, master_sheet.max_row + 1):
        year = master_sheet["C" + str(i)].value
        next_col = master_sheet["C" + str(i + 1)].value
        if year == next_col:
            year_sheet = file[str(year)]
            year_sheet["B" + str(row)] = master_sheet["A" + str(i)].value
            year_sheet["C" + str(row)] = master_sheet["B" + str(i)].value
            year_sheet["D" + str(row)] = master_sheet["C" + str(i)].value
            year_sheet["E" + str(row)] = master_sheet["D" + str(i)].value
            year_sheet["F" + str(row)] = master_sheet["E" + str(i)].value
            year_sheet["G" + str(row)] = master_sheet["F" + str(i)].value
            year_sheet["H" + str(row)] = master_sheet["G" + str(i)].value
            year_sheet["I" + str(row)] = master_sheet["H" + str(i)].value
            year_sheet["J" + str(row)] = master_sheet["I" + str(i)].value
            year_sheet["K" + str(row)] = master_sheet["J" + str(i)].value

            row+=1
        else:
            row = 2

    
#setting the headers 
#row 1
def header(file, master_sheet):
    for i in range(2, master_sheet.max_row):
        #if the cohort year changes, move to the next sheet
        year = master_sheet["C" + str(i)].value
        next_col = master_sheet["C" + str(i + 1)].value
        year_sheet = file[str(year)]
        if year != next_col:
            year_sheet["B1"] = "First Name"
            year_sheet["C1"] = "Last Name"
            year_sheet["D1"] = "Year"
            year_sheet["E1"] = "Email"
            year_sheet["F1"] = "Major"
            year_sheet["G1"] = "State"
            year_sheet["H1"] = "City"
            year_sheet["I1"] = "Position"
            year_sheet["J1"] = "Company"
            year_sheet["K1"] = "Last Updated"

def main():

    alum_file = load_workbook(filename="Alumni Records.xlsx")
    # print(alum_file.sheetnames)
    #Gets the top sheet and it's title
    master = input("What file are you sorting?: ")
    master_sheet = alum_file[master]

    header(alum_file, master_sheet)
    sort(alum_file, master_sheet)

    alum_file.save(filename="Alumni Records.xlsx")
    print("done")
    #identifying headers
  
main()