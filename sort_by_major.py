from openpyxl import Workbook, load_workbook, workbook
from openpyxl.worksheet import worksheet
from alum import Alum

def importing(file, master_sheet):
    #writing all the people into a list
    #assign each into the alum class
    list_of_alums = []
    row = 2
    for i in range(2, master_sheet.max_row + 1):
        major = master_sheet["F" + str(i)].value
        alumni = Alum()
        alumni.first_name = master_sheet["A" + str(i)].value
        alumni.last_name = master_sheet["B" + str(i)].value
        alumni.cohort_year = master_sheet["C" + str(i)].value
        alumni.email = master_sheet["D" + str(i)].value
        alumni.major = master_sheet["E" + str(i)].value
        alumni.state = master_sheet["F" + str(i)].value
        alumni.city = master_sheet["G" + str(i)].value
        alumni.position = master_sheet["H" + str(i)].value
        alumni.company = master_sheet["I" + str(i)].value
        alumni.update = master_sheet["J" + str(i)].value

        list_of_alums.append()

def sort(file, alum_list):
    for alumus in alum_list:
        if alumus.major not in file.sheetnames:
            major_sheet= file.create_sheet(title=alumus.major)
        


        
    
#setting the headers 
#row 1
def header(file, master_sheet):
    for i in range(2, master_sheet.max_row):
        #if the cohort year changes, move to the next sheet
        year = master_sheet["C" + str(i)].value
        next_col = master_sheet["C" + str(i + 1)].value
        year_sheet = file[str(year)]
        if year != next_col:
            year_sheet["B1"] = "First Name"
            year_sheet["C1"] = "Last Name"
            year_sheet["D1"] = "Year"
            year_sheet["E1"] = "Email"
            year_sheet["F1"] = "Major"
            year_sheet["G1"] = "State"
            year_sheet["H1"] = "City"
            year_sheet["I1"] = "Position"
            year_sheet["J1"] = "Company"
            year_sheet["K1"] = "Last Updated"

def main():

    alum_file = load_workbook(filename="Alumni Records.xlsx")
    # print(alum_file.sheetnames)
    #Gets the top sheet and it's title
    master = input("What file are you sorting?: ")
    master_sheet = alum_file[master]

    header(alum_file, master_sheet)
    sort(alum_file, master_sheet)

    alum_file.save(filename="Alumni Records.xlsx")
    print("done")
    #identifying headers
  
main()